"""
Step 1: Import the grants spreadsheet into a SQLite database.

Usage:
    python import_grants.py [path/to/grants.xlsx]

Reads the "Grants Database" sheet, parses deadlines into machine-readable
dates where possible, and writes everything to grants.db.
Safe to re-run: it rebuilds the table each time (source of truth = xlsx,
until the scrapers take over that job later).
"""

import re
import sqlite3
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

DB_PATH = Path(__file__).parent / "grants.db"
DEFAULT_XLSX = Path(__file__).parent / "data" / "grants.xlsx"

MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

REGIONS = [
    ("austria", "Austria"), ("germany", "Germany"), ("switzerland", "Switzerland"),
    ("swiss", "Switzerland"), ("united kingdom", "UK"), ("uk", "UK"),
    ("ireland", "Ireland"), ("denmark", "Denmark"), ("sweden", "Sweden"),
    ("norway", "Norway"), ("netherlands", "Netherlands"), ("france", "France"),
    ("spain", "Spain"), ("portugal", "Portugal"), ("italy", "Italy"),
    ("israel", "Israel"), ("japan", "Japan"), ("canada", "Canada"),
    ("poland", "Poland"), ("belgium", "Belgium"), ("finland", "Finland"),
    ("czech", "Czechia"), ("hungary", "Hungary"), ("slovakia", "Slovakia"),
    ("flanders", "Belgium"),
    ("australia", "Australia"), ("new zealand", "New Zealand"), ("singapore", "Singapore"),
    ("hong kong", "Hong Kong"), ("korea", "South Korea"), ("china", "China"),
    ("taiwan", "Taiwan"), ("india", "India"),
    ("usa", "USA"), ("united states", "USA"),
    ("us ", "USA"), ("embc", "EU"), ("european", "EU"), ("eu ", "EU"),
    ("eu+", "EU"), ("international", "International"), ("worldwide", "International"),
    ("global", "International"),
]

def normalize_region(text):
    """Map messy region text to one clean primary label.
    'Austria -> USA' => Austria (earliest country named wins)."""
    if not text:
        return "International"
    t = " " + text.lower() + " "
    best, best_pos = None, 10**9
    for key, label in REGIONS:
        pos = t.find(key)
        if pos != -1 and pos < best_pos:
            best, best_pos = label, pos
    return best or "International"


SENIOR_PD = ["advanced postdoc", "senior postdoc", "pre-professorship",
             "group leader", "independence", "pathway to independence"]
ESTABLISHED = ["faculty", "established", "professor", "principal investigator",
               "investigator", "tenure", "independent researcher", "teams",
               "senior researcher", "chairs", "all career stages"]

CAT_RULES = [
    ("Startups & Innovation", ["startup", "start-up", "spin-off", "commercialisation",
                               "incubat", "accelerat", "venture", "founding", "founders",
                               "sme", "innovation"]),
    ("Life Sciences & Medicine", ["biomed", "health", "life science", "medicin", "medical", "biotech"]),
    ("Natural Sciences & Engineering", ["natural science", "stem", "engineering", "physic", "math",
                                        "technical", "deep-tech", "science & engineering", "ict", "environment"]),
    ("Social Sciences & Humanities", ["humanit", "social", "cultural", "econom", "philosoph", "religion"]),
    ("Mobility & Career", ["mobility", "career", "doctoral training", "knowledge transfer", "research stays"]),
]

def normalize_category(text):
    """Map free-text discipline to general 'major' buckets (semicolon-joined)."""
    t = (text or "").lower()
    buckets = [label for label, keys in CAT_RULES if any(k in t for k in keys)]
    # generic / cross-disciplinary signals => All disciplines
    if (not [b for b in buckets if b != "Mobility & Career"]) or any(
            k in t for k in ["all discipline", "all areas", "all main", "all science",
                             "any", "basic research", "frontier", "curiosity",
                             "interdisciplinary", "high-risk", "novel ideas", "big questions"]):
        buckets.append("All disciplines")
    seen, out = set(), []
    for b in buckets:
        if b not in seen:
            seen.add(b); out.append(b)
    return ";".join(out)


BRIDGE_KEYS = ["spin-off", "spinoff", "transition", "proof of concept", "academic",
               "r&d", "research", "bridge", "dissertation"]

def classify_audience(category, title, description, career, startup_stage=""):
    """Label each grant: For research | For startups | For startups doing research.

    A filled-in Startup Stage column is the definitive signal that an entry is
    startup funding, regardless of how its free-text category is worded.
    """
    cat = normalize_category(category)
    blob = " ".join([category or "", title or "", description or "", career or ""]).lower()
    if startup_stage.strip() or "Startups & Innovation" in cat:
        if any(k in blob for k in BRIDGE_KEYS):
            return "For startups doing research"
        return "For startups"
    return "For research"


def normalize_career(text):
    """Map free-text career stage to general buckets (semicolon-joined)."""
    t = (text or "").lower()
    buckets = []
    if "master" in t:
        buckets.append("Master")
    # strip postdoc terms so "postDOCTORAL" doesn't false-match the PhD check
    t_nopd = t.replace("post-doctoral", "").replace("postdoctoral", "").replace("post-doc", "").replace("postdoc", "").replace("post-phd", "").replace("post phd", "")
    if any(k in t_nopd for k in ["phd", "doctoral", "doctorate", "dphil", "graduate student"]):
        buckets.append("PhD")
    if any(k in t for k in SENIOR_PD):
        buckets.append("Senior Postdoc")
        if "early" in t:
            buckets.append("Early Postdoc")
    elif "postdoc" in t or "post-doc" in t:
        if "early" in t:
            buckets.append("Early Postdoc")
        else:  # generic "postdoctoral" fits both
            buckets += ["Early Postdoc", "Senior Postdoc"]
    if not buckets and ("early-career" in t or "early career" in t):
        buckets.append("Early Postdoc")
    if any(k in t for k in ESTABLISHED) or "any" in t or not buckets:
        buckets.append("Any / Established")
    seen, out = set(), []
    for b in buckets:
        if b not in seen:
            seen.add(b); out.append(b)
    return ";".join(out)


def parse_deadline(text):
    """
    Best-effort parse of the free-text deadline column.

    Returns (deadline_type, iso_date or None):
      - 'rolling'   : continuous submission, no date
      - 'fixed'     : a concrete date was found  -> iso_date set
      - 'approx'    : only month+year found      -> iso_date = 1st of month
      - 'check_site': recurring/unclear          -> no date
    """
    if not text:
        return "check_site", None
    t = text.lower()

    if "rolling" in t or "continuous" in t or "no fixed deadline" in t:
        return "rolling", None

    date_pat = r"\b(\d{1,2})\s+(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+(\d{4})"

    # If the text names an explicit deadline (e.g. "opens 3 Aug 2026; deadline 3 Nov 2026"),
    # prefer the date that follows the word "deadline" over an earlier opening date.
    m = re.search(r"deadline\D{0,10}" + date_pat, t)
    if not m:
        # Otherwise take the first full date: "27 Aug 2026", "9 Sep 2026"
        m = re.search(date_pat, t)
    if m:
        day, mon, year = int(m.group(1)), MONTHS[m.group(2)], int(m.group(3))
        try:
            return "fixed", date(year, mon, day).isoformat()
        except ValueError:
            pass

    # Month + year only: "Oct 2026", "Est. Jan 2027"
    m = re.search(r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.?\s+(\d{4})", t)
    if m:
        mon, year = MONTHS[m.group(1)], int(m.group(2))
        return "approx", date(year, mon, 1).isoformat()

    return "check_site", None


def _table_exists(con, name):
    return con.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (name,)
    ).fetchone() is not None


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        sys.exit(f"Spreadsheet not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Grants Database"]

    rows = ws.iter_rows(values_only=True)
    header = None
    records = []
    for row in rows:
        cells = ["" if c is None else str(c).strip() for c in row]
        # The header row starts with "ID"
        if header is None:
            if cells and cells[0] == "ID":
                header = cells
            continue
        if not cells[0]:  # skip empty rows
            continue
        records.append(cells)

    if header is None:
        sys.exit("Could not find the header row (expected a row starting with 'ID').")

    col = {name: i for i, name in enumerate(header)}

    def get(rec, name):
        i = col.get(name)
        return rec[i] if i is not None and i < len(rec) else ""

    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    # Preserve auto-ingested rows across re-imports (skip if old schema)
    try:
        if _table_exists(con, "grants"):
            cur.execute("CREATE TABLE IF NOT EXISTS _keep AS SELECT * FROM grants WHERE origin != 'curated'")
    except sqlite3.OperationalError:
        pass  # old schema without origin column — nothing to preserve
    cur.execute("DROP TABLE IF EXISTS _keep_invalid")
    cur.execute("DROP TABLE IF EXISTS grants")
    cur.execute("""
        CREATE TABLE grants (
            id            TEXT PRIMARY KEY,
            title         TEXT NOT NULL,
            funding_body  TEXT,
            source        TEXT,
            region        TEXT,
            category      TEXT,
            career_stage  TEXT,
            description   TEXT,
            amount        TEXT,
            duration      TEXT,
            eligibility   TEXT,
            deadline_text TEXT,   -- raw text from the sheet
            deadline_date TEXT,   -- parsed ISO date (nullable)
            deadline_type TEXT,   -- fixed | approx | rolling | check_site
            application_mode TEXT,
            interview     TEXT,
            success_rate  TEXT,
            link          TEXT,
            notes         TEXT,
            region_clean  TEXT DEFAULT '',
            career_clean  TEXT DEFAULT '',
            category_clean TEXT DEFAULT '',
            audience      TEXT DEFAULT 'For research',
            startup_stage  TEXT DEFAULT '',
            company_status TEXT DEFAULT '',
            funding_type   TEXT DEFAULT '',
            origin        TEXT DEFAULT 'curated',  -- curated | grants_gov | eu_portal
            last_checked  TEXT DEFAULT (date('now'))
        )
    """)

    for rec in records:
        deadline_text = get(rec, "Next Deadline (2026/27)")
        dtype, ddate = parse_deadline(deadline_text)
        cur.execute(
            "INSERT INTO grants VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'curated',date('now'))",
            (
                get(rec, "ID"), get(rec, "Title"), get(rec, "Funding Body"),
                get(rec, "Source"), get(rec, "Country / Region"), get(rec, "Category"),
                get(rec, "Career Stage"), get(rec, "Description"), get(rec, "Amount (funding)"),
                get(rec, "Duration"), get(rec, "Eligibility"),
                deadline_text, ddate, dtype,
                get(rec, "Application Mode"), get(rec, "Interview Stage?"),
                get(rec, "Approx. Success Rate"), get(rec, "Link"), get(rec, "Notes"),
                normalize_region(get(rec, "Country / Region")),
                normalize_career(get(rec, "Career Stage")),
                (normalize_category(get(rec, "Category"))
                 + (";Startups & Innovation"
                    if get(rec, "Startup Stage").strip()
                    and "Startups & Innovation" not in normalize_category(get(rec, "Category"))
                    else "")),
                classify_audience(get(rec, "Category"), get(rec, "Title"),
                                  get(rec, "Description"), get(rec, "Career Stage"),
                                  get(rec, "Startup Stage")),
                get(rec, "Startup Stage"), get(rec, "Company Status"),
                get(rec, "Funding Type"),
            ),
        )

    # restore auto-ingested rows
    if _table_exists(con, "_keep"):
        try:
            cur.execute("INSERT OR IGNORE INTO grants SELECT * FROM _keep")
        except sqlite3.OperationalError:
            pass  # schema changed; auto rows will be re-ingested
        cur.execute("DROP TABLE _keep")

    con.commit()
    n = cur.execute("SELECT COUNT(*) FROM grants").fetchone()[0]
    by_type = cur.execute(
        "SELECT deadline_type, COUNT(*) FROM grants GROUP BY deadline_type"
    ).fetchall()
    con.close()

    print(f"Imported {n} grants into {DB_PATH}")
    for t, c in by_type:
        print(f"  deadline_type={t}: {c}")


if __name__ == "__main__":
    main()
